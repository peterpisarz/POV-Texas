import openpyxl
import json

# Load the Excel workbook
wb = openpyxl.load_workbook('data.xlsx')

# Select the active sheet
ws = wb.active

# Initialize an empty list to store the data
data = []

# Get the number of rows in the sheet
num_rows = ws.max_row

# Define the column indices for each field
name_col = 1
desc_col = 2
trait_col = 3
value_col = 4

# Iterate through each row in the sheet
for row in range(2, num_rows + 1):
    # Append the data as a dictionary to the list
    data.append({
        'name': ws.cell(row, name_col).value,
        'description': ws.cell(row, desc_col).value,
        'trait_type': ws.cell(row, trait_col).value,
        'value': ws.cell(row, value_col).value
    })

# Convert the data to JSON
json_data = json.dumps(data, indent=2, separators=(',', ': '))
json_data = json_data.replace("}, {", "},\n{")

# Write the JSON data to a file
with open('data.json', 'w') as f:
    f.write(json_data)
